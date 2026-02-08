import pandas as pd
import os
import re
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlrd # Explicit import for PyInstaller hidden import
import unicodedata

# Suppress openpyxl warnings if any
warnings.filterwarnings("ignore")

def process_files(input_files, output_file, target_year=None, attendance_file=None):
    """
    Reads selected Excel files, aggregates comments, and saves to output_file.
    Filters by target_year if provided (checks Column C).
    Sorts by attendance_file if provided (Col B=ID, Col C=Name).
    """
    all_data = []
    
    print(f"Processing {len(input_files)} files...")
    print(f"Target Year Filter: {target_year if target_year else 'None'}")
    
    attendance_ids_order = [] # List of IDs in order
    attendance_map_name = {}  # ID -> Name (Kanji)
    attendance_map_original_id = {} # NormID -> Original ID (for display)
    attendance_name_map = {}  # Normalized Name (Kanji) -> ID (for fuzzy match)
    if attendance_file:
        print(f"Loading attendance sheet: {os.path.basename(attendance_file)}")
        try:
            # Read Attendance Sheet
            # Assume Col B (Index 1) = ID, Col C (Index 2) = Name
            # User specified: Data starts from Row 7 (index 6). Rows 1-6 are junk.
            att_df = pd.read_excel(attendance_file, header=None)
            
            # Slice to skip first 6 rows
            if len(att_df) > 6:
                att_df = att_df.iloc[6:]
            else:
                print("Attendance sheet has fewer than 7 rows.")
                return False, "Attendance sheet is too short/empty."
            
            for _, row in att_df.iterrows():
                if pd.isna(row[1]):
                    continue
                
                s_id = str(row[1]).strip()
                s_name = str(row[2]).strip() if not pd.isna(row[2]) else ""
                
                # Heuristic to skip headers
                if s_id.lower() in ["学籍番号", "id", "student id", "headerid", "number"]:
                    continue
                
                # Store by normalized ID (lowercase)
                norm_id = s_id.lower()
                attendance_ids_order.append(norm_id)
                attendance_map_name[norm_id] = s_name
                attendance_map_original_id[norm_id] = s_id # Store original
                
                # Normalize Name for Fuzzy Matching
                # NFKC converts full-width to half-width
                # Then remove all whitespaces
                norm_name = unicodedata.normalize('NFKC', s_name)
                norm_name = re.sub(r'\s+', '', norm_name)
                if norm_name:
                    attendance_name_map[norm_name] = norm_id
                
            print(f"Loaded {len(attendance_ids_order)} students from attendance sheet.")
            
        except Exception as e:
            print(f"Error reading attendance sheet: {e}")
            return False, f"Error reading attendance sheet: {e}"

    # Map to store comment colors: (NormID, Date) -> PatternFill object or color hex
    # Storing hex string is easier for re-creating PatternFill
    comment_colors = {} 

    for file_path in input_files:
        try:
            filename = os.path.basename(file_path)
            date_match = re.search(r'\d{4}-\d{2}-\d{2}', filename)
            date_str = date_match.group(0) if date_match else filename
            
            # Use openpyxl to read data AND styles
            wb_in = load_workbook(file_path, data_only=True) # data_only=True gets values, but styles are on the cell
            ws_in = wb_in.active
            
            # Convert to list of rows to iterate easily
            # We expect columns A-G (0-6)
            # Row 1 might be header, but we usually assume raw data as per previous logic (header=None in pandas)
            # Pandas read_excel(header=None) treats 1st row as data. 
            # So we iterate all rows.
            
            for row in ws_in.iter_rows():
                # Check sufficient columns (at least 7: A..G)
                if len(row) < 7:
                    continue
                    
                # Extract values (converting to string same as pandas default roughly)
                def get_val(cell):
                    return str(cell.value) if cell.value is not None else ""
                
                sub_id_col = get_val(row[0])
                course_col = get_val(row[2])
                name_col   = get_val(row[4])
                id_col     = get_val(row[5])
                comment_col = get_val(row[6])
                
                # Get Style from Comment Cell (row[6])
                comment_cell = row[6]
                fill_color = None
                if comment_cell.fill and comment_cell.fill.patternType == 'solid':
                    # Extract ARGB hex
                    fg = comment_cell.fill.start_color
                    if fg.type == 'rgb':
                        fill_color = fg.rgb # e.g. "FFFF0000"
                    elif fg.type == 'theme':
                         # Theme colors are hard to resolve without theme map.
                         # Often ignored, or we assume mapped manually?
                         # For now, stick to RGB if available.
                         # Many simple highlights are RGB.
                         pass
                    
                    # Store color if it looks like a valid highlight (not white/transparent)
                    # Default '00000000' or similar might appear.
                    if fill_color and fill_color not in ['00000000', 'FFFFFFFF', '00FFFFFF']:
                        pass
                    else:
                        fill_color = None

                if not id_col.strip(): 
                    if not name_col.strip(): 
                        continue
                
                if target_year and not course_col.strip().startswith(str(target_year)):
                    continue
                
                try:
                    sub_id_val = float(sub_id_col)
                except ValueError:
                    sub_id_val = 0.0

                # 1. Primary Match: Normalize ID
                norm_id = id_col.strip().lower()
                
                # 2. Check if ID exists in Attendance
                if norm_id in attendance_map_name:
                    final_name = attendance_map_name[norm_id]
                else:
                    # 3. Fallback: Fuzzy Name Match
                    # Normalize Comment Name
                    c_name_norm = unicodedata.normalize('NFKC', name_col.strip())
                    c_name_norm = re.sub(r'\s+', '', c_name_norm)
                    
                    found_id = None
                    for att_name_norm, att_id in attendance_name_map.items():
                        if c_name_norm.startswith(att_name_norm):
                            found_id = att_id
                            break
                    
                    if found_id:
                        norm_id = found_id
                        final_name = attendance_map_name[found_id]
                    else:
                        final_name = name_col.strip()
                
                # Store Data
                all_data.append({
                    'SubmissionID': sub_id_val,
                    'Name': final_name,
                    'ID': id_col.strip(),
                    'NormID': norm_id,
                    'Date': date_str,
                    'Comment': comment_col.strip()
                })
                
                # Store Color if exists
                if fill_color:
                    # Key by (NormID, Date)
                    # Note: Duplicates handled later, but we just store latest color for now is fine
                    # Or store in all_data and pick during pivot?
                    # Storing in a separate dict is tricky if multiple submissions.
                    # Bester approach: add 'Color' to all_data, let deduplication handle it.
                    all_data[-1]['Color'] = fill_color

        except Exception as e:
            print(f"Error processing {file_path}: {e}")
            return False, f"Error processing {os.path.basename(file_path)}: {str(e)}"

    if not all_data and not attendance_ids_order:
        return False, "No data found."

    full_df = pd.DataFrame(all_data)
    
    # --- Deduplication Logic (same as before, but on NormID + Date) ---
    if not full_df.empty:
        full_df = full_df.sort_values(by=['NormID', 'Date', 'SubmissionID'], ascending=[True, True, True])
        full_df = full_df.drop_duplicates(subset=['NormID', 'Date'], keep='last')
    
    # Extract Color Map after deduplication
    color_map = {} # (Name, Date) or (NormID, Date) -> ColorHex
                   # Wait, pivot uses NormID.
    if not full_df.empty and 'Color' in full_df.columns:
         for _, row in full_df.iterrows():
             if not pd.isna(row['Color']):
                 color_map[(row['NormID'], row['Date'])] = row['Color']

    # Pivot
    if not full_df.empty:
        pivot_df = full_df.pivot_table(
            index=['NormID'], 
            columns='Date', 
            values='Comment', 
            aggfunc='last'
        )
    else:
        pivot_df = pd.DataFrame()

    # --- Sorting / Reindexing by ID ---
    if attendance_ids_order:
        # Reindex by Normalized ID
        existing_ids = pivot_df.index.tolist() if not pivot_df.empty else []
        extra_ids = [i for i in existing_ids if i not in attendance_ids_order]
        full_order_ids = attendance_ids_order + extra_ids
        
        if not pivot_df.empty:
            pivot_df = pivot_df.reindex(full_order_ids)
        else:
            pivot_df = pd.DataFrame(index=full_order_ids)
            
        pivot_df.index.name = 'NormID'
        pivot_df = pivot_df.reset_index()
        
        # Restore Display Columns (Name, ID)
        extras_map = {}
        if not full_df.empty:
            for _, row in full_df.iterrows():
                extras_map[row['NormID']] = {'Name': row['Name'], 'ID': row['ID']}
        
        final_names = []
        final_ids = []
        
        for nid in pivot_df['NormID']:
            if nid in attendance_map_name:
                final_names.append(attendance_map_name[nid]) 
            elif nid in extras_map:
                final_names.append(extras_map[nid]['Name']) 
            else:
                final_names.append("Unknown") 
                
            # ID (Display)
            # Priority: Attendance Original ID > Comment Sheet Original ID > NormID
            if nid in attendance_map_original_id:
                final_ids.append(attendance_map_original_id[nid])
            elif nid in extras_map:
                final_ids.append(extras_map[nid]['ID'])
            else:
                 final_ids.append(nid) 
                 
        pivot_df.insert(0, 'Name', final_names)
        pivot_df.insert(1, 'ID', final_ids)
        
    else:
        # No attendance sheet
        if not full_df.empty:
            pivot_df = pivot_df.reset_index() 
            meta_df = full_df[['NormID', 'Name', 'ID']].drop_duplicates(subset=['NormID'])
            pivot_df = pd.merge(pivot_df, meta_df, on='NormID', how='left')
        else:
             return False, "No data found."
             
    # Fill NaN
    pivot_df = pivot_df.fillna("未回答")
    
    # Columns
    if 'NormID' in pivot_df.columns:
        # Keep NormID for Styling Map Lookup!
        # pivot_df['NormID'] used for lookup
        pass 
    
    # Reorder columns
    cols = list(pivot_df.columns)
    # NormID might be there. We need it for row-lookup during styling.
    # But we shouldn't save it to file if we can avoid it.
    # Or strict mapping: NormID is preserved in memory or we reconstruct map using (ID, Name)?
    # Name/ID might be ambiguous? NormID is unique key.
    # Let's keep NormID in DF, save without it, then reload? 
    # Or iterate DF to apply style before saving?
    # Pandas to_excel does not allow easy styling per cell during write.
    # We have to write, then load with openpyxl.
    # If we drop NormID, we lose the key.
    # But wait, output rows correspond to `pivot_df` rows exactly.
    # We can iterate `pivot_df` rows and apply style to Excel rows by index.
    
    # Final Columns for Output
    # We want Name, ID, [Dates...]
    # We should exclude NormID and Color from OUTPUT but use them for index tracking.
    dates = [c for c in cols if c not in ['Name', 'ID', 'NormID']]
    dates.sort()
    output_cols = ['Name', 'ID'] + dates
    
    # Create final DF for saving
    save_df = pivot_df[output_cols]
    
    print(f"Saving summary to {output_file}")
    try:
        save_df.to_excel(output_file, index=False)
        
        # 5. Styling
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Define Fills
        # Light Red for Unanswered
        fill_unanswered = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Iterate rows
        # Row 1 is header. Data starts row 2.
        # pivot_df index 0 corresponds to Excel Row 2.
        
        for idx, row_series in pivot_df.iterrows():
            excel_row_idx = idx + 2
            norm_id = row_series['NormID'] if 'NormID' in row_series else None
            
            # Iterate date columns
            for col_offset, date_col in enumerate(dates):
                # Columns: Name(1), ID(2), Date1(3), Date2(4)...
                excel_col_idx = 3 + col_offset
                cell = ws.cell(row=excel_row_idx, column=excel_col_idx)
                
                val = str(cell.value)
                
                if val == "未回答":
                    cell.fill = fill_unanswered
                else:
                    # Check for preserved color
                    if norm_id:
                        color_hex = color_map.get((norm_id, date_col))
                        if color_hex:
                            try:
                                cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                            except:
                                pass # formatting error
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50:
                adjusted_width = 50
            ws.column_dimensions[column].width = adjusted_width
            
        wb.save(output_file)
        print("Done.")
        return True, f"Saved to {os.path.basename(output_file)}"
    except PermissionError:
        return False, f"Permission denied: {output_file}. Close it and try again."
    except Exception as e:
        return False, f"Error saving file: {str(e)}"
